import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime


def pre_excel(excel_path, output_path, label="请评价样品"):
    # 删除前6列
    df = pd.read_excel(excel_path).iloc[:, 6:]
    df.rename(columns={df.columns[0]: "姓名"}, inplace=True)
    # 找到  请评价样品  列
    specified_indices = []
    for index, col in enumerate(df.columns):
        if label in col:
            specified_indices.append(index)

    # 初始化一个空字典，用于存储拆分后的DataFrame
    sheets = {}

    # 遍历指定的列索引，将列索引之间的DataFrame逐个放到新的sheet中
    for i in range(len(specified_indices) - 1):
        start_idx = specified_indices[i]
        end_idx = specified_indices[i + 1]
        sheet_name = f"Sheet_{i + 1}"
        # 加上原始sheet的第一列
        sheets[sheet_name] = df.iloc[:, [0] + list(range(start_idx, end_idx))]

    # 将最后一个指定列索引之后的DataFrame放到新的sheet中
    sheet_name = f"Sheet_{len(specified_indices)}"
    sheets[sheet_name] = df.iloc[
        :, [0] + list(range(specified_indices[-1], len(df.columns)))
    ]

    # 将拆分后的DataFrame写入不同的sheet
    with pd.ExcelWriter(output_path) as writer:
        for sheet_name, sheet_df in sheets.items():
            # 1. 从第3列开始后的所有列重命名，取列名中文顿号后的字符串
            new_columns = sheet_df.columns[:2].tolist() + [
                col.split("、")[1] if "、" in col else col
                for col in sheet_df.columns[2:]
            ]
            # 2. 第2列取"请评价样品"和后续"样品"之间的字符串作为该sheet的新名称
            second_col_name = sheet_df.columns[1]
            new_columns[1] = second_col_name.split("—")[1]
            sheet_df.columns = new_columns
            new_sheet_name = second_col_name.split(label)[1].split("样品")[0].strip()
            # 按照人名排序
            sheet_df_sorted = sheet_df.sort_values(by=sheet_df.columns[0])
            sheet_df_sorted.insert(1, "样品", new_sheet_name)
            # 写入新的sheet
            sheet_df_sorted.to_excel(writer, sheet_name=new_sheet_name, index=False)


def calculate_fp(cost_limit, sheet_map, sheet_code_map):
    excel_files = []
    for file in os.listdir():
        if file.endswith(".xlsx") and "属性评分" in file:
            excel_files.append(os.path.join(os.getcwd(), file))

    tmp_file1 = "temp_output1.xlsx"
    tmp_file2 = "temp_output2.xlsx"
    pre_excel(excel_files[0], tmp_file1)
    pre_excel(excel_files[1], tmp_file2)

    with pd.ExcelFile(tmp_file1) as xls1, pd.ExcelFile(tmp_file2) as xls2:
        sheet_names = xls1.sheet_names
        # sheet_map = {386: 385, 472: 826, 631: 170, 902: 759}
        # sheet_code_map = {386: "Anchor", 472: "TATUA", 631: "PRESIDENT", 902: "Debic"}

        now = datetime.now()
        current_summary = now.strftime("%Y-%m-%d_%H_%M_%S")
        output_file = f"{current_summary}_summary.xlsx"

        # # 读取所有sheet
        # xls1 = pd.ExcelFile(tmp_file1)
        # xls2 = pd.ExcelFile(tmp_file2)

        # # 获取所有sheet名称
        # sheet_names = xls1.sheet_names
        # sheet_map = {386: 385, 472: 826, 631: 170, 902: 759}
        # sheet_code_map = {386: "Anchor", 472: "TATUA", 631: "PRESIDENT", 902: "Debic"}

        # # 创建一个工作簿对象，用于保存修改后的Excel文件
        # now = datetime.now()
        # current_summary = now.strftime("%Y-%m-%d_%H_%M_%S")
        # output_file = f"{current_summary}_summary.xlsx"
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet in sheet_names:
                # 读取当前sheet的数据
                df1 = pd.read_excel(xls1, sheet_name=sheet)
                df2 = pd.read_excel(xls2, sheet_name=str(sheet_map[int(sheet)]))

                # 插入数据
                df1 = pd.concat([df1, df2], ignore_index=True)

                df1_sorted = df1.sort_values(by=[df1.columns[0], df1.columns[1]])

                # 保存到新的Excel文件
                df1_sorted.to_excel(
                    writer, sheet_name=sheet_code_map[int(sheet)], index=False
                )
    # 加载修改后的Excel文件
    wb = load_workbook(output_file)
    ws = wb.active

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # 遍历每两行
        for row in range(2, ws.max_row + 1, 2):
            # 统计大于等于cost_limit的个数
            count = 0
            for col in range(3, ws.max_column + 1):
                cell1 = ws.cell(row=row, column=col)
                cell2 = ws.cell(row=row + 1, column=col)

                # 计算差值
                diff = abs(cell1.value - cell2.value)

                # 标记大于等于cost_limit的单元格
                if diff >= cost_limit:
                    count += 1
                    cell1.fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )
                    cell2.fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )
            # 标记人名单元格
            if count >= 4:
                ws.cell(row=row, column=1).fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                ws.cell(row=row + 1, column=1).fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                print(
                    f"针对---->{sheet}<----样品的测试, {ws.cell(row=row, column=1).value}分数不达标"
                )
    # 保存修改后的Excel文件
    wb.save(output_file)
    wb.close()
    # 删除临时文件
    if os.path.isfile(tmp_file1):
        os.remove(tmp_file1)
    if os.path.isfile(tmp_file2):
        os.remove(tmp_file2)


if __name__ == "__main__":
    if os.path.isfile("config.xlsx"):
        df = pd.read_excel("config.xlsx")
        sheet_map, sheet_code_map = {}, {}
        for _, row in df.iterrows():
            sheet_map[row.to_dict()["code1"]] = row.to_dict()["code2"]
            sheet_code_map[row.to_dict()["code1"]] = row.to_dict()["name"]
    else:
        sheet_map={386: 385, 472: 826, 631: 170, 902: 759}
        sheet_code_map={386: "Anchor", 472: "TATUA", 631: "PRESIDENT", 902: "Debic"}
    cost_limit = float(input("请输入分差阈值: "))
    calculate_fp(cost_limit=cost_limit, sheet_map=sheet_map, sheet_code_map=sheet_code_map)
    input("请按任意键结束: ")
